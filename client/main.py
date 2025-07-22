import argparse
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 1. command line arguments

parser = argparse.ArgumentParser(description="CSV upload and Excel generation")
parser.add_argument("-k", "--keys", nargs="+", help="Extra fields to include (e.g., kurzname info)", default=[])
parser.add_argument("-c", "--colored", action="store_true", help="Enable row coloring based on 'hu'")
args = parser.parse_args()

# 2.vehiclses.csv file to the API
with open("vehicles.csv", "rb") as f:
    response = requests.post("http://127.0.0.1:8000/api/upload/", files={"file": f})

# 3. error control
if response.status_code != 200:
    print("API error:", response.text)
    exit(1)

# 4. successful response
data = response.json()

# 5. Create Excel 
wb = Workbook()
ws = wb.active
ws.title = "Vehicles"

# 6. header row
headers = ["rnr", "gruppe"] + args.keys
ws.append(headers)

# 7. Hex colors for row coloring
COLOR_GREEN = PatternFill(fgColor="007500", fill_type="solid")
COLOR_ORANGE = PatternFill(fgColor="FFA500", fill_type="solid")
COLOR_RED = PatternFill(fgColor="b30000", fill_type="solid")

# 8. all items in the response
for item in sorted(data, key=lambda x: x.get("gruppe", "")):
    row = [item.get("rnr", ""), item.get("gruppe", "")]
    
    for key in args.keys:
        row.append(item.get(key, ""))

    ws.append(row)
    row_index = ws.max_row

    # 9. coloring rows based on 'hu' date
    if args.colored:
        hu_date = item.get("hu")
        if hu_date:
            try:
                hu = datetime.fromisoformat(hu_date)
                diff_months = (datetime.today().year - hu.year) * 12 + (datetime.today().month - hu.month)
                if diff_months <= 3:
                    fill = COLOR_GREEN
                elif diff_months <= 12:
                    fill = COLOR_ORANGE
                else:
                    fill = COLOR_RED

                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_index, column=col).fill = fill
            except Exception as e:
                print("Tarih işlenemedi:", e)

# 10. save the Excel file
today = datetime.today().date().isoformat()
filename = f"vehicles_{today}.xlsx"
wb.save(filename)
print(f"Excel kaydedildi: {filename}")
